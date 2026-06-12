"""
chunk4.py – Docling-only RFP / Document Chunker  (No LLM)
==========================================================
Supported formats (all via Docling):
  .pdf  .docx  .xlsx  .pptx  .html  .htm  .md  .csv
  .txt  .asciidoc  .latex  .email

Key design decisions
  • Zero LLM calls  – all metadata comes from Docling's document model
  • Max chunk size  ≈ 200 tokens  (TARGET_MAX_CHARS = 800, ~4 chars/token)
  • After converting each file, the full DoclingDocument pydantic model
    summary is printed before chunking begins
  • Table-aware splitting: column-header row stored in table_header and
    prepended to every sub-chunk
  • XLSX sheet/tab names extracted from the DoclingDocument page labels
  • Text paragraphs that share the same heading are merged up to
    TARGET_MAX_CHARS, then split at sentence boundaries
  • Plain .txt falls back to paragraph-based native chunking if Docling
    does not recognise it as a richer format
"""

import csv
import gc
import json
import logging
import os
import re
import tempfile
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

import fitz  # PyMuPDF  (PDF page-batching only)

# ── Logging ───────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)-8s | %(message)s",
)
logger = logging.getLogger(__name__)

# ── Docling ───────────────────────────────────────────────────────────────────
project_models = os.path.join(os.path.dirname(os.path.abspath(__file__)), "models")
easyocr_path   = os.path.join(project_models, "EasyOcr")
os.environ["EASYOCR_MODULE_PATH"] = easyocr_path

from docling.chunking import HierarchicalChunker                         # noqa: E402
from docling.datamodel.base_models import InputFormat                    # noqa: E402
from docling.datamodel.pipeline_options import (                         # noqa: E402
    AcceleratorDevice,
    AcceleratorOptions,
    EasyOcrOptions,
    PdfPipelineOptions,
    TableFormerMode,
)
from docling.document_converter import DocumentConverter, PdfFormatOption  # noqa: E402

# ═══════════════════════════════════════════════════════════════════════════════
# CONFIGURATION
# ═══════════════════════════════════════════════════════════════════════════════

TARGET_MAX_CHARS     = 800   # ≈ 200 tokens  (4 chars / token avg for English)
TARGET_MIN_CHARS     = 100    # merge chunks shorter than this with neighbours
TABLE_ROWS_PER_CHUNK = 5     # data rows per table sub-chunk  (header excluded)
PDF_BATCH_SIZE       = 5     # pages per PDF batch

# Absolute hard-cap: any chunk (table or text) beyond this is force-split
HARD_MAX_CHARS       = TARGET_MAX_CHARS * 2   # 1600 chars — safety net

input_folder = Path("./RFP_TEST/SBC")
output_file  = "F_chunked_rfp_data.json"

# ── Extension → InputFormat  (for Docling-handled formats) ───────────────────
DOCLING_FORMAT_MAP: Dict[str, InputFormat] = {
    ".pdf":       InputFormat.PDF,
    ".docx":      InputFormat.DOCX,
    ".xlsx":      InputFormat.XLSX,
    ".pptx":      InputFormat.PPTX,
    ".html":      InputFormat.HTML,
    ".htm":       InputFormat.HTML,
    ".md":        InputFormat.MD,
    ".csv":       InputFormat.CSV,
    ".asciidoc":  InputFormat.ASCIIDOC,
    ".adoc":      InputFormat.ASCIIDOC,
    ".latex":     InputFormat.LATEX,
    ".tex":       InputFormat.LATEX,
    ".eml":       InputFormat.EMAIL,
}

# Formats handled natively (fallback if Docling can't load them)
NATIVE_TXT_FORMATS = {".txt"}

# ── Docling converter setup ───────────────────────────────────────────────────
_pdf_opts = PdfPipelineOptions()
_pdf_opts.do_ocr             = True
_pdf_opts.do_table_structure = True
_pdf_opts.artifacts_path     = project_models
_pdf_opts.images_scale       = 1.0
_pdf_opts.table_structure_options.mode = TableFormerMode.ACCURATE
_pdf_opts.ocr_options = EasyOcrOptions(
    lang=["en"],
    model_storage_directory=easyocr_path,
    download_enabled=False,
)
_pdf_opts.accelerator_options = AcceleratorOptions(device=AcceleratorDevice.CPU)

converter = DocumentConverter(
    allowed_formats=list(DOCLING_FORMAT_MAP.values()),
    format_options={
        InputFormat.PDF: PdfFormatOption(pipeline_options=_pdf_opts),
    },
)

chunker = HierarchicalChunker()

# ═══════════════════════════════════════════════════════════════════════════════
# DOCLING DOCUMENT PYDANTIC SUMMARY PRINTER
# ═══════════════════════════════════════════════════════════════════════════════

def print_doc_summary(doc, filename: str) -> None:
    """
    Print a readable summary of the DoclingDocument pydantic model
    immediately after conversion, before chunking begins.

    Shows:
      * Page / sheet count
      * Element-type counts  (text, table, picture, ...)
      * First few headings found
      * Table dimensions
      * Top-level pydantic model fields
    """
    bar = "-" * 64
    try:
        print(f"\n{bar}")
        print(f"  DOCLING DOCUMENT MODEL  >>  {filename}")
        print(bar)
    except Exception:
        return  # terminal can't even print ASCII -- give up silently

    try:
        # -- page count ------------------------------------------------
        pages = getattr(doc, "pages", {})
        if isinstance(pages, dict):
            print(f"  Pages / sheets   : {len(pages)}")
        elif hasattr(pages, "__len__"):
            print(f"  Pages / sheets   : {len(pages)}")

        # -- element counts -------------------------------------------
        def _safe_len(attr):
            obj = getattr(doc, attr, None)
            if obj is None:
                return 0
            if isinstance(obj, dict):
                return len(obj)
            try:
                return len(obj)
            except Exception:
                return "?"

        elem_counts = {
            "texts":    _safe_len("texts"),
            "tables":   _safe_len("tables"),
            "pictures": _safe_len("pictures"),
            "groups":   _safe_len("groups"),
        }
        for k, v in elem_counts.items():
            if v:
                print(f"  {k:<16s} : {v}")

        # -- top-level pydantic fields --------------------------------
        try:
            fields = list(doc.model_fields.keys())
            print(f"  Model fields     : {fields}")
        except Exception:
            pass

        # -- first few section headings -------------------------------
        headings_found: List[str] = []
        texts = getattr(doc, "texts", None) or []
        for item in texts:
            lbl = getattr(item, "label", None)
            lbl_str = (lbl.value if hasattr(lbl, "value") else str(lbl)) if lbl else ""
            if "heading" in lbl_str.lower() or "title" in lbl_str.lower():
                text_val = getattr(item, "text", "") or ""
                if text_val.strip():
                    headings_found.append(text_val.strip()[:80])
            if len(headings_found) >= 6:
                break
        if headings_found:
            print(f"  First headings   :")
            for h in headings_found:
                try:
                    print(f"    * {h}")
                except UnicodeEncodeError:
                    print(f"    * [non-printable heading]")

        # -- table dimensions ----------------------------------------
        tables = getattr(doc, "tables", None) or []
        if tables:
            print(f"  Tables ({len(tables)})        :")
            for i, tbl in enumerate(list(tables)[:6]):
                data = getattr(tbl, "data", None)
                rows = getattr(data, "num_rows", "?") if data else "?"
                cols = getattr(data, "num_cols", "?") if data else "?"
                cap  = ""
                captions = getattr(tbl, "captions", None)
                if captions:
                    try:
                        cap = f"  [{str(list(captions)[0])[:40]}]"
                    except Exception:
                        pass
                print(f"    [{i}] {rows} rows x {cols} cols{cap}")

        # -- pydantic JSON snippet (first 400 chars) ------------------
        try:
            snippet = doc.model_dump_json(indent=2)[:400]
            print(f"\n  Pydantic JSON (first 400 chars):")
            for ln in snippet.split("\n"):
                try:
                    print(f"    {ln}")
                except UnicodeEncodeError:
                    print(f"    [line contains non-printable chars]")
            print("    ...")
        except Exception as exc:
            print(f"  [JSON dump skipped: {exc}]")

    except Exception as exc:
        print(f"  [Summary error: {exc}]")

    print(bar + "\n")



# ═══════════════════════════════════════════════════════════════════════════════
# CHUNK RECORD FACTORY
# ═══════════════════════════════════════════════════════════════════════════════

def _make_chunk(
    content:           str,
    chunk_type:        str,           # "table" | "text" | "list" | "native_txt" | "native_csv"
    source_file:       str,
    doc_format:        str,           # file extension  e.g. ".xlsx"
    headings:          List[str],
    page_numbers:      List[int],
    element_types:     List[str],
    sheet_name:        str            = "",
    table_header:      str            = "",
    table_chunk_index: int            = 0,
    subheadings:       List[str]      = None,
    captions:          List[str]      = None,
    page_range:        str            = "",
    section_depth:     int            = 0,
) -> Dict[str, Any]:
    """
    Central factory for all chunk records.
    chunk_index is assigned to 0 here and replaced in the finalize step.
    word_count and char_count are computed from content.
    """
    words = len(content.split())
    return {
        "content":        content,
        "content_length": len(content),
        "word_count":     words,
        "created_at":     datetime.now().isoformat(),
        "chunk_type":     chunk_type,
        "metadata": {
            "source_file":       source_file,
            "doc_format":        doc_format,
            "chunk_index":       0,
            "headings":          headings,
            "subheadings":       subheadings or [],
            "captions":          captions    or [],
            "element_types":     element_types,
            "page_numbers":      page_numbers,
            "sheet_name":        sheet_name,
            "table_header":      table_header,
            "table_chunk_index": table_chunk_index,
            "page_range":        page_range,
            "section_depth":     section_depth,
        },
    }


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET-NAME EXTRACTION  (from DoclingDocument pages for XLSX)
# ═══════════════════════════════════════════════════════════════════════════════

def _build_sheet_map(doc, file_path: Path) -> Dict[int, str]:
    """
    Build a  {page_no (1-indexed): sheet_name}  mapping for XLSX files.

    Strategy (in priority order):
      1. Check doc.pages – some Docling versions expose a page_label field.
      2. Fall back to openpyxl to read sheetnames by index.
      3. Return empty dict if neither works.
    """
    sheet_map: Dict[int, str] = {}

    # ── Strategy 1: DoclingDocument pages ────────────────────────
    pages = getattr(doc, "pages", {})
    pg_items = pages.values() if isinstance(pages, dict) else (pages or [])
    for pg in pg_items:
        pg_no  = getattr(pg, "page_no", None)
        label  = getattr(pg, "page_label", None)
        if pg_no is not None and label:
            sheet_map[int(pg_no)] = str(label)

    if sheet_map:
        return sheet_map

    # ── Strategy 2: openpyxl fallback ────────────────────────────
    if file_path.suffix.lower() == ".xlsx":
        try:
            import openpyxl
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            sheet_map = {i + 1: name for i, name in enumerate(wb.sheetnames)}
            wb.close()
        except Exception as exc:
            logger.warning(f"  openpyxl sheet-name fallback failed: {exc}")

    return sheet_map


# ═══════════════════════════════════════════════════════════════════════════════
# STRUCTURAL METADATA EXTRACTION  (from a single Docling chunk)
# ═══════════════════════════════════════════════════════════════════════════════

_HEADING_RE = re.compile(
    r"^(?:[A-Z]\.|[IVX]+\.|[0-9]+\.(?:[0-9]+\.)*|#+)\s+.{3,}$",
    re.MULTILINE,
)


def _extract_meta(chunk, sheet_map: Dict[int, str] = None) -> Dict[str, Any]:
    """
    Extract all structural metadata from a Docling HierarchicalChunk.
    Returns a dict with:
      headings, subheadings, captions, element_types, page_numbers,
      sheet_name, section_depth
    """
    meta: Dict[str, Any] = {
        "headings":      [],
        "subheadings":   [],
        "captions":      [],
        "element_types": [],
        "page_numbers":  [],
        "sheet_name":    "",
        "section_depth": 0,
    }

    cm = getattr(chunk, "meta", None)
    if cm is None:
        return meta

    # ── headings ──────────────────────────────────────────────────
    if hasattr(cm, "headings") and cm.headings:
        meta["headings"] = [str(h) for h in cm.headings if h]
        meta["section_depth"] = len(meta["headings"])

    # ── captions ──────────────────────────────────────────────────
    if hasattr(cm, "captions") and cm.captions:
        meta["captions"] = [str(c) for c in cm.captions if c]

    # ── doc_items: element labels + page numbers ──────────────────
    if hasattr(cm, "doc_items") and cm.doc_items:
        labels: set = set()
        pages:  set = set()
        for item in cm.doc_items:
            # label
            lbl = getattr(item, "label", None)
            if lbl is not None:
                labels.add(lbl.value if hasattr(lbl, "value") else str(lbl))

            # page / sheet number
            for prov in (getattr(item, "prov", None) or []):
                pg = getattr(prov, "page_no", None) or getattr(prov, "page", None)
                if pg is not None:
                    pages.add(int(pg))

        meta["element_types"] = sorted(labels)
        meta["page_numbers"]  = sorted(pages)

        # resolve sheet name for XLSX
        if sheet_map and meta["page_numbers"]:
            first_pg  = meta["page_numbers"][0]
            meta["sheet_name"] = sheet_map.get(first_pg, "")

    # ── fallback subheadings from text pattern ────────────────────
    if not meta["headings"]:
        raw = (
            getattr(chunk, "text", None)
            or getattr(cm, "text", None)
            or ""
        )
        if raw:
            found = _HEADING_RE.findall(raw)
            meta["subheadings"] = [h.strip() for h in found if len(h.strip()) < 120]

    return meta


# ═══════════════════════════════════════════════════════════════════════════════
# SENTENCE-BOUNDARY SPLITTING
# ═══════════════════════════════════════════════════════════════════════════════

def _split_sentences(text: str, max_chars: int = TARGET_MAX_CHARS) -> List[str]:
    """
    Break text into pieces of at most max_chars, splitting at sentence
    boundaries (`.  !  ?`) where possible.
    """
    if len(text) <= max_chars:
        return [text]

    sentences = re.split(r"(?<=[.!?])\s+", text)
    parts: List[str] = []
    cur = ""
    for sent in sentences:
        if not cur:
            cur = sent
        elif len(cur) + 1 + len(sent) <= max_chars:
            cur += " " + sent
        else:
            if cur:
                parts.append(cur)
            if len(sent) > max_chars:
                # hard-split very long single sentence
                for i in range(0, len(sent), max_chars):
                    parts.append(sent[i : i + max_chars])
                cur = ""
            else:
                cur = sent
    if cur:
        parts.append(cur)
    return [p for p in parts if p.strip()]


# ═══════════════════════════════════════════════════════════════════════════════
# TABLE SPLITTING  (for Docling table chunks)
# ═══════════════════════════════════════════════════════════════════════════════

def _extract_table_header_from_docling(chunk) -> str:
    """
    Extract the true column-header row directly from the DoclingDocument
    TableItem data model (grid cells, row 0).
    Falls back to empty string if unavailable.
    This preserves the first-column label and all other column labels.
    """
    cm = getattr(chunk, "meta", None)
    if cm is None:
        return ""
    for item in (getattr(cm, "doc_items", None) or []):
        data = getattr(item, "data", None)
        if data is None:
            continue
        grid = getattr(data, "grid", None)
        if not grid:
            continue
        # First row of the grid — collect all cell texts
        header_cells = []
        first_row = grid[0] if grid else []
        for cell in first_row:
            txt = ""
            spans = getattr(cell, "spans", None)
            if spans:
                txt = str(spans[0]).strip() if spans else ""
            if not txt:
                txt = str(getattr(cell, "text", "") or "").strip()
            if txt:
                header_cells.append(txt)
        if header_cells:
            return " | ".join(header_cells[:20])[:300]
    return ""


def _parse_table_rows(text: str, headings: List[str]) -> tuple:
    """
    Return (header_text, [row_strings]) by parsing the Docling-serialised
    table text.

    Handles four shapes:
      A – newline rows with a header line that has no '='
      B – flat 'Sec = Val. Sec = Val' rows (key=value style)
      C – repeated-heading VRM style  (heading repeats every row)
      D – fallback: return all lines as data rows
    """
    working = text.strip()

    # Strip leading section headings
    for h in headings:
        h_stripped = h.strip()
        if h_stripped and working.startswith(h_stripped):
            working = working[len(h_stripped):].lstrip("\n :")

    lines = [ln.strip() for ln in working.split("\n") if ln.strip()]
    if not lines:
        return "", [working]

    # ── Shape A: multi-line, first line is header (no '=') ──────────
    if len(lines) >= 2:
        first = lines[0]
        # Candidate header: has multiple pipe/dot-separated tokens, no '='
        if " = " not in first:
            cells = [c.strip() for c in re.split(r"[|.]+", first) if c.strip()]
            if len(cells) >= 2:
                header_text = " | ".join(cells)
                return header_text, lines[1:]

    # ── Shape B: key = value flat rows ──────────────────────────────
    # Detect by counting occurrences of ' = '
    eq_count = working.count(" = ")
    if eq_count >= 2:
        # Split on period-space that precedes a capital or comma
        parts = [p.strip().rstrip(".") for p in
                 re.split(r"\.\s+(?=[A-Z0-9,])", working) if p.strip()]
        if parts:
            # First part without '=' is the header candidate
            if " = " not in parts[0]:
                return parts[0][:300], parts[1:]
            return "", parts

    # ── Shape C / D: fallback – treat all lines as data rows ────────
    return "", lines


def _hard_split_content(
    content:       str,
    header_text:   str,
    headings:      List[str],
    source_file:   str,
    doc_format:    str,
    page_numbers:  List[int],
    page_range:    str,
    captions:      List[str],
    element_types: List[str],
    sheet_name:    str,
    section_depth: int,
    chunk_type:    str = "table",
) -> List[Dict[str, Any]]:
    """
    Absolute last-resort splitter: splits content at sentence/period
    boundaries, prepending the table header + section headings to every
    sub-chunk so context is preserved.  Guarantees chunks <= HARD_MAX_CHARS.
    """
    heading_prefix = "\n".join(headings) if headings else ""
    prefix_parts: List[str] = []
    if heading_prefix:
        prefix_parts.append(heading_prefix)
    if header_text:
        prefix_parts.append(f"Columns: {header_text}")
    prefix = "\n".join(prefix_parts)
    prefix_len = len(prefix) + (1 if prefix else 0)  # +1 for the newline join

    available = HARD_MAX_CHARS - prefix_len
    if available < 100:
        available = HARD_MAX_CHARS  # fallback: ignore prefix in size calc

    # Split at sentence boundaries
    pieces = _split_sentences(content, max_chars=available)

    result = []
    for idx, piece in enumerate(pieces):
        if not piece.strip():
            continue
        body_parts = []
        if heading_prefix:
            body_parts.append(heading_prefix)
        if header_text:
            body_parts.append(f"Columns: {header_text}")
        body_parts.append(piece)
        result.append(_make_chunk(
            content="\n".join(body_parts),
            chunk_type=chunk_type,
            source_file=source_file,
            doc_format=doc_format,
            headings=headings,
            page_numbers=page_numbers,
            element_types=element_types,
            captions=captions,
            sheet_name=sheet_name,
            table_header=header_text,
            table_chunk_index=idx,
            page_range=page_range,
            section_depth=section_depth,
        ))
    return result


def _split_table(
    content:       str,
    headings:      List[str],
    source_file:   str,
    doc_format:    str,
    page_numbers:  List[int],
    page_range:    str,
    captions:      List[str],
    element_types: List[str],
    sheet_name:    str          = "",
    section_depth: int          = 0,
    header_text:   str          = "",   # pre-extracted from docling grid
) -> List[Dict[str, Any]]:
    """
    Split an oversized Docling table chunk into sub-chunks of
    TABLE_ROWS_PER_CHUNK rows each.  The column-header row is prepended
    to every sub-chunk and stored in metadata.table_header.

    Falls back to _hard_split_content() if row parsing produces too few
    rows or rows that are still oversized.
    """
    heading_prefix = "\n".join(headings) if headings else ""

    # Use pre-extracted docling grid header first; fall back to parser
    parsed_header, data_rows = _parse_table_rows(content, headings)
    if not header_text:
        header_text = parsed_header

    # ── Immediate fallback: no rows parsed OR rows are huge ───────
    usable_rows = [r for r in data_rows if r.strip()]
    if not usable_rows or any(len(r) > HARD_MAX_CHARS for r in usable_rows):
        return _hard_split_content(
            content=content,
            header_text=header_text,
            headings=headings,
            source_file=source_file,
            doc_format=doc_format,
            page_numbers=page_numbers,
            page_range=page_range,
            captions=captions,
            element_types=element_types,
            sheet_name=sheet_name,
            section_depth=section_depth,
        )

    def _sub(batch: List[str], idx: int) -> Dict[str, Any]:
        parts = []
        if heading_prefix:
            parts.append(heading_prefix)
        if header_text:
            parts.append(f"Columns: {header_text}")
        parts.append(". ".join(batch))
        sub_content = "\n".join(parts)
        # If this sub-chunk is still too big, hard-split it
        if len(sub_content) > HARD_MAX_CHARS:
            return None  # signal for hard-split
        return _make_chunk(
            content=sub_content,
            chunk_type="table",
            source_file=source_file,
            doc_format=doc_format,
            headings=headings,
            page_numbers=page_numbers,
            element_types=element_types,
            captions=captions,
            sheet_name=sheet_name,
            table_header=header_text,
            table_chunk_index=idx,
            page_range=page_range,
            section_depth=section_depth,
        )

    result: List[Dict[str, Any]] = []
    chunk_idx = 0
    for start in range(0, len(usable_rows), TABLE_ROWS_PER_CHUNK):
        batch = usable_rows[start : start + TABLE_ROWS_PER_CHUNK]
        c = _sub(batch, chunk_idx)
        if c is None:
            # batch is still huge → hard-split batch text
            batch_text = ". ".join(batch)
            sub_chunks = _hard_split_content(
                content=batch_text,
                header_text=header_text,
                headings=headings,
                source_file=source_file,
                doc_format=doc_format,
                page_numbers=page_numbers,
                page_range=page_range,
                captions=captions,
                element_types=element_types,
                sheet_name=sheet_name,
                section_depth=section_depth,
            )
            result.extend(sub_chunks)
            chunk_idx += len(sub_chunks)
        elif c["content"].strip():
            result.append(c)
            chunk_idx += 1

    if not result:
        # Ultimate fallback
        return _hard_split_content(
            content=content,
            header_text=header_text,
            headings=headings,
            source_file=source_file,
            doc_format=doc_format,
            page_numbers=page_numbers,
            page_range=page_range,
            captions=captions,
            element_types=element_types,
            sheet_name=sheet_name,
            section_depth=section_depth,
        )
    return result


# ═══════════════════════════════════════════════════════════════════════════════
# TEXT MERGING  (for Docling text chunks)
# ═══════════════════════════════════════════════════════════════════════════════

def _merge_text(
    raw:        List[Dict[str, Any]],
    doc_format: str,
) -> List[Dict[str, Any]]:
    """
    Merge consecutive text chunks that share the same primary heading
    until TARGET_MAX_CHARS is reached, then flush and split as needed.
    """
    if not raw:
        return []

    result: List[Dict[str, Any]] = []
    buf: Dict[str, Any] = {
        "content": "", "headings": [], "subheadings": [], "captions": [],
        "pages": [], "elems": set(), "page_range": "", "source": "",
        "sheet": "", "depth": 0,
    }

    def _flush() -> None:
        if not buf["content"].strip():
            return
        for piece in _split_sentences(buf["content"].strip()):
            if piece.strip():
                result.append(_make_chunk(
                    content=piece,
                    chunk_type="text",
                    source_file=buf["source"],
                    doc_format=doc_format,
                    headings=buf["headings"][:],
                    subheadings=buf["subheadings"][:],
                    captions=buf["captions"][:],
                    page_numbers=sorted(set(buf["pages"])),
                    element_types=sorted(buf["elems"]),
                    sheet_name=buf["sheet"],
                    page_range=buf["page_range"],
                    section_depth=buf["depth"],
                ))
        buf["content"] = ""

    def _reset(c: Dict[str, Any]) -> None:
        m = c["metadata"]
        buf["content"]    = c["content"]
        buf["headings"]   = m.get("headings",      [])[:]
        buf["subheadings"]= m.get("subheadings",   [])[:]
        buf["captions"]   = m.get("captions",      [])[:]
        buf["pages"]      = m.get("page_numbers",  [])[:]
        buf["elems"]      = set(m.get("element_types", []))
        buf["page_range"] = m.get("page_range",    "")
        buf["source"]     = m["source_file"]
        buf["sheet"]      = m.get("sheet_name",    "")
        buf["depth"]      = m.get("section_depth", 0)

    for chunk in raw:
        m       = chunk["metadata"]
        content = chunk["content"]
        h0      = (m.get("headings") or [""])[0]

        if not buf["content"]:
            _reset(chunk)
            continue

        same_head = (h0 == (buf["headings"] or [""])[0] or not h0)
        fits      = len(buf["content"]) + 1 + len(content) <= TARGET_MAX_CHARS

        if same_head and fits:
            buf["content"] += " " + content
            buf["pages"].extend(m.get("page_numbers", []))
            buf["elems"].update(m.get("element_types", []))
            if m.get("page_range"):
                buf["page_range"] = m["page_range"]
        else:
            _flush()
            _reset(chunk)

    _flush()
    return result


# ═══════════════════════════════════════════════════════════════════════════════
# CORE DOCLING CHUNK PIPELINE
# ═══════════════════════════════════════════════════════════════════════════════

def _pipeline(
    raw_chunks: List,
    source_file: str,
    doc_format:  str,
    page_range:  str              = "",
    sheet_map:   Dict[int, str]   = None,
) -> List[Dict[str, Any]]:
    """
    Convert a list of raw Docling HierarchicalChunks into our records.
    Tables → split if oversized.
    Text   → queue for merging.
    Order is preserved.
    """
    sheet_map = sheet_map or {}
    result:       List[Dict[str, Any]] = []
    pending_text: List[Dict[str, Any]] = []

    def _flush_text() -> None:
        if pending_text:
            result.extend(_merge_text(pending_text, doc_format))
            pending_text.clear()

    for chunk in raw_chunks:
        text = chunker.contextualize(chunk)
        if not text or not text.strip():
            continue

        meta     = _extract_meta(chunk, sheet_map)
        is_table = any(t in meta["element_types"] for t in ("table", "document_index"))
        content  = text.strip()

        if is_table:
            _flush_text()

            # ── Extract column header from docling grid (most accurate) ──
            # Preserves first-column label + all other column names.
            tbl_header = _extract_table_header_from_docling(chunk)

            # Fallback: parse from contextualized text
            if not tbl_header:
                lines = content.split("\n")
                if lines:
                    # Find first line that is NOT a section heading
                    for ln in lines:
                        ln_s = ln.strip()
                        if ln_s and ln_s not in meta["headings"]:
                            # Take it as header only if it looks columnar
                            # (multiple pipe/period/comma-separated tokens, no '=')
                            if " = " not in ln_s:
                                cells = [c.strip() for c in
                                         re.split(r"[|,.]+", ln_s) if c.strip()]
                                if len(cells) >= 2:
                                    tbl_header = " | ".join(cells[:20])[:300]
                            break

            if len(content) <= TARGET_MAX_CHARS:
                result.append(_make_chunk(
                    content=content,
                    chunk_type="table",
                    source_file=source_file,
                    doc_format=doc_format,
                    headings=meta["headings"],
                    subheadings=meta["subheadings"],
                    captions=meta["captions"],
                    page_numbers=meta["page_numbers"],
                    element_types=meta["element_types"],
                    sheet_name=meta["sheet_name"],
                    table_header=tbl_header,
                    table_chunk_index=0,
                    page_range=page_range,
                    section_depth=meta["section_depth"],
                ))
            else:
                result.extend(_split_table(
                    content=content,
                    headings=meta["headings"],
                    source_file=source_file,
                    doc_format=doc_format,
                    page_numbers=meta["page_numbers"],
                    page_range=page_range,
                    captions=meta["captions"],
                    element_types=meta["element_types"],
                    sheet_name=meta["sheet_name"],
                    section_depth=meta["section_depth"],
                    header_text=tbl_header,   # pass pre-extracted header
                ))
        else:
            # ── Hard-cap oversized text chunks too ──────────────────────
            if len(content) > HARD_MAX_CHARS:
                for piece in _split_sentences(content, max_chars=TARGET_MAX_CHARS):
                    if piece.strip():
                        pending_text.append(_make_chunk(
                            content=piece,
                            chunk_type="text",
                            source_file=source_file,
                            doc_format=doc_format,
                            headings=meta["headings"],
                            subheadings=meta["subheadings"],
                            captions=meta["captions"],
                            page_numbers=meta["page_numbers"],
                            element_types=meta["element_types"],
                            sheet_name=meta["sheet_name"],
                            page_range=page_range,
                            section_depth=meta["section_depth"],
                        ))
            else:
                pending_text.append(_make_chunk(
                    content=content,
                    chunk_type="text",
                    source_file=source_file,
                    doc_format=doc_format,
                    headings=meta["headings"],
                    subheadings=meta["subheadings"],
                    captions=meta["captions"],
                    page_numbers=meta["page_numbers"],
                    element_types=meta["element_types"],
                    sheet_name=meta["sheet_name"],
                    page_range=page_range,
                    section_depth=meta["section_depth"],
                ))

    _flush_text()
    return result


# ═══════════════════════════════════════════════════════════════════════════════
# FORMAT-SPECIFIC PROCESSORS
# ═══════════════════════════════════════════════════════════════════════════════

def _fitz_page_text(pdf_doc, page_idx: int) -> str:
    """
    Extract plain text from a single PDF page via PyMuPDF.
    Used as fallback when Docling yields no chunks for a page.
    """
    try:
        page = pdf_doc[page_idx]
        return page.get_text("text").strip()
    except Exception:
        return ""


def process_pdf(file_path: Path, source_file_name: str) -> List[Dict[str, Any]]:
    """
    PDF: page-by-page batching via PyMuPDF + Docling.

    Extra robustness:
      • First batch gets a fitz raw-text fallback so page-1 cover text
        is never silently dropped.
      • Each batch chunk count is compared to a fitz text presence check;
        if docling returned nothing for a batch that clearly has text,
        we emit a native_txt chunk as a safety net.
    """
    all_chunks: List[Dict[str, Any]] = []
    pdf_doc     = fitz.open(file_path)
    total_pages = len(pdf_doc)
    print(f"    Total pages: {total_pages}")

    for start in range(0, total_pages, PDF_BATCH_SIZE):
        end            = min(start + PDF_BATCH_SIZE - 1, total_pages - 1)
        page_range_str = f"{start + 1}-{end + 1}"
        try:
            tmp_doc = fitz.open()
            tmp_doc.insert_pdf(pdf_doc, from_page=start, to_page=end)
            with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
                tmp_path = tmp.name
            tmp_doc.save(tmp_path)
            tmp_doc.close()

            print(f"      Batch pages {start+1}-{end+1}")
            result     = converter.convert(tmp_path)
            doc        = result.document

            # Print pydantic summary for first batch only
            if start == 0:
                try:
                    print_doc_summary(doc, source_file_name)
                except Exception:
                    pass  # never let a print failure abort the pipeline

            raw_chunks = list(chunker.chunk(doc))
            print(f"         -> {len(raw_chunks)} raw Docling chunks")

            batch_chunks = _pipeline(
                raw_chunks,
                source_file=source_file_name,
                doc_format=".pdf",
                page_range=page_range_str,
            )

            # ── Fallback: if docling returned nothing, use fitz text ──
            if not batch_chunks:
                fitz_texts = []
                for pg_idx in range(start, end + 1):
                    pg_text = _fitz_page_text(pdf_doc, pg_idx)
                    if pg_text:
                        fitz_texts.append(pg_text)
                combined = " ".join(fitz_texts).strip()
                if combined:
                    print(f"         -> (fitz fallback for pages {start+1}-{end+1})")
                    for piece in _split_sentences(combined, max_chars=TARGET_MAX_CHARS):
                        if piece.strip():
                            batch_chunks.append(_make_chunk(
                                content=piece,
                                chunk_type="text",
                                source_file=source_file_name,
                                doc_format=".pdf",
                                headings=[],
                                page_numbers=list(range(start + 1, end + 2)),
                                element_types=["text"],
                                page_range=page_range_str,
                            ))

            all_chunks.extend(batch_chunks)
            os.remove(tmp_path)
            del result, raw_chunks
            gc.collect()

        except Exception as exc:
            print(f"      Failed batch {start+1}-{end+1}: {exc}")
            import traceback; traceback.print_exc()

    pdf_doc.close()
    return all_chunks


def process_with_docling(
    file_path:        Path,
    source_file_name: str,
    fmt:              InputFormat,
) -> List[Dict[str, Any]]:
    """
    Generic Docling processor for DOCX, XLSX, PPTX, HTML, MD, CSV, etc.
    Prints the pydantic summary after conversion.
    """
    try:
        result = converter.convert(file_path)
        doc    = result.document

        try:
            print_doc_summary(doc, source_file_name)
        except Exception:
            pass  # never let a print failure abort the pipeline

        raw_chunks = list(chunker.chunk(doc))
        print(f"    -> {len(raw_chunks)} raw Docling chunks")

        doc_format = file_path.suffix.lower()
        sheet_map  = _build_sheet_map(doc, file_path) if doc_format == ".xlsx" else {}

        if sheet_map:
            print(f"    Sheet names: {list(sheet_map.values())}")

        chunks = _pipeline(
            raw_chunks,
            source_file=source_file_name,
            doc_format=doc_format,
            sheet_map=sheet_map,
        )
        del result, raw_chunks
        gc.collect()
        return chunks

    except Exception as exc:
        print(f"    Error: {exc}")
        import traceback; traceback.print_exc()
        return []


def process_txt(file_path: Path, source_file_name: str) -> List[Dict[str, Any]]:
    """
    Native plain-text processor.
    Splits on blank lines (paragraph boundaries) and then at sentence
    boundaries to enforce TARGET_MAX_CHARS.
    """
    print(f"    Processing as plain text (native)")
    try:
        text = file_path.read_text(encoding="utf-8", errors="replace")
    except Exception as exc:
        print(f"    Read error: {exc}")
        return []

    # Split into paragraphs
    paragraphs = [p.strip() for p in re.split(r"\n{2,}", text) if p.strip()]
    print(f"    -> {len(paragraphs)} paragraphs found")

    chunks: List[Dict[str, Any]] = []
    for para in paragraphs:
        for piece in _split_sentences(para):
            if piece.strip():
                chunks.append(_make_chunk(
                    content=piece,
                    chunk_type="native_txt",
                    source_file=source_file_name,
                    doc_format=".txt",
                    headings=[],
                    page_numbers=[],
                    element_types=["text"],
                ))
    return chunks


# ═══════════════════════════════════════════════════════════════════════════════
# FINALIZE  (assign global chunk indices, word counts)
# ═══════════════════════════════════════════════════════════════════════════════

def _finalize(chunks: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    for i, c in enumerate(chunks):
        c["metadata"]["chunk_index"] = i
        c["content_length"]          = len(c["content"])
        c["word_count"]              = len(c["content"].split())
    return chunks


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════════

def chunk_single_document(file_path: str, session_id: str = "") -> List[Dict[str, Any]]:
    """
    Parse and chunk a single document file.
    Returns a list of chunk dictionaries with rich metadata.
    """
    path = Path(file_path)
    suffix = path.suffix.lower()

    if suffix not in DOCLING_FORMAT_MAP and suffix not in NATIVE_TXT_FORMATS:
        raise ValueError(f"Unsupported format: {suffix}")

    try:
        if suffix == ".pdf":
            chunks = process_pdf(path, path.name)
        elif suffix in NATIVE_TXT_FORMATS:
            chunks = process_txt(path, path.name)
        else:
            fmt    = DOCLING_FORMAT_MAP[suffix]
            chunks = process_with_docling(path, path.name, fmt)

        chunks = _finalize(chunks)
        
        # Inject session_id if provided
        if session_id:
            for c in chunks:
                c["metadata"]["session_id"] = session_id
                
        return chunks

    except Exception as exc:
        logger.error(f"Failed to chunk {file_path}: {exc}")
        import traceback; traceback.print_exc()
        raise

def chunk_multiple_documents(file_paths: list, session_id: str = "") -> list:
    """Processes a list of files (Called by live uploads)"""
    master_chunk_list = []
    for file_path in file_paths:
        try:
            print(f"Chunking file: {file_path}")
            chunks = chunk_single_document(file_path, session_id)
            master_chunk_list.extend(chunks)
        except Exception as exc:
            print(f"[ERROR] Skipping {file_path}: {exc}")
    return master_chunk_list

def chunk_directory(folder_path: str) -> list:
    """Processes an entire recursive folder (Called once to build historical DB)"""
    all_chunks = []
    input_folder = Path(folder_path)
    for file_path in input_folder.rglob("*"):
        if file_path.is_dir(): continue
        try:
            all_chunks.extend(chunk_single_document(str(file_path)))
        except Exception as exc:
            pass # Skip unsupported or errored files silently for batch processing
    return all_chunks
